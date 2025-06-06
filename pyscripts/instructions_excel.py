from qgis.PyQt import uic
import os
from qgis.PyQt.QtWidgets import QDialog, QFileDialog, QTableWidgetItem, QMessageBox
import xlrd
import openpyxl
import re

# Load the .ui file
FORM_CLASS, _ = uic.loadUiType(
    os.path.join(
        os.path.dirname(__file__), '..', 'UI', 'InstructionsExcel.ui'))


class InstructionsExcel(QDialog, FORM_CLASS):
    """A dialog for loading data from an Excel file."""

    def __init__(
            self, table_widget, layer_name, line_edit1, line_edit2, line_edit3,
            line_edit4, push_button_1, push_button_2, push_button_3,
            push_button_4, main_instance, delete_polygon, clear_all_rows,
            numbering, distance_calc, area_calc, create_polygonMA,
            create_polygonPA):
        """Initialize the InstructionsExcel dialog."""

        super().__init__()

        # Set up the user interface from Designer
        self.setupUi(self)

        # Make the dialog modal
        self.setModal(True)

        # Store
        self.table_widget = table_widget
        self.layer_name = layer_name
        self.line_edit1 = line_edit1
        self.line_edit2 = line_edit2
        self.line_edit3 = line_edit3
        self.line_edit4 = line_edit4
        self.push_button_1 = push_button_1
        self.push_button_2 = push_button_2
        self.push_button_3 = push_button_3
        self.push_button_4 = push_button_4
        self.main_instance = main_instance
        self.delete_polygon = delete_polygon
        self.clear_all_rows = clear_all_rows
        self.numbering = numbering
        self.distance_calc = distance_calc
        self.area_calc = area_calc
        self.create_polygonMA = create_polygonMA
        self.create_polygonPA = create_polygonPA

        self.pushButton.clicked.connect(self.closeEvent)
        self.pushButton_2.clicked.connect(self.load_excel)

    def get_language_key(self):
        label_text = self.label_21.text()
        if label_text == 'Օրինակ՝':
            return 'hy'  # Armenian
        elif label_text == 'Example:':
            return 'en'  # English
        elif label_text == 'Ejemplo:':
            return 'es'  # Spanish
        return 'en'  # Default to English

    def closeEvent(self, event):
        """Handle the close event."""

        self.close()

    def load_excel(self):
        """Load data from an Excel file."""

        lang = self.get_language_key()

        messages = {
            'select_excel': {
                'hy': 'Ընտրել համապատասխան Excel ֆայլը',
                'en': 'Select the corresponding Excel file',
                'es': 'Seleccionar el archivo Excel correspondiente'
            },
            'insufficient_data': {
                'hy': 'Տվյալները բացակայում են կամ բավարար չեն։\nExcel ֆայլը '\
                    'պետք է պարունակի տեղեկատվություն առաջին երեք '\
                    'սյունակներում:',
                'en': 'Missing or insufficient data.\nThe Excel file must '\
                    'contain information in the first three columns.',
                'es': 'Faltan datos o son insuficientes.\nEl fichero Excel '\
                    'debe contener información en las tres primeras columnas.'
            },
            'invalid_format': {
                'hy': 'Առաջին սյունակի տվյալների ձևաչափը վավեր չէ:',
                'en': 'The format of data in the first column is not valid.',
                'es': 'El formato de los datos de la primera columna no es válido.'
            },
            'min_points_required': {
                'hy': 'Գրանցման համար անհրաժեշտ է առնվազն 3 կետ։',
                'en': 'A minimum of 3 points is required to be registered.',
                'es': 'Se requiere un mínimo de 3 puntos para ser registrado.'
            },
            'warning_title': {
                'hy': 'Ուշադրություն',
                'en': 'Warning',
                'es': 'Atención'
            }
        }

        title = messages['select_excel'][lang]
        msg_box = QMessageBox()
        msg_box.information(
            self, messages['warning_title'][lang],
            messages['select_excel'][lang])

        path, _ = QFileDialog.getOpenFileName(
            self, title, '', 'Excel (*.xlsx *xls)')
        if path:
            # Clear old data
            if self.table_widget.rowCount() != 1:
                numRows = self.table_widget.rowCount()
                for row in range(1, numRows):
                    self.table_widget.removeRow(1)
                self.delete_polygon(
                    self.main_instance, self.layer_name)
                if self.layer_name == 'Monument_Area':
                    self.line_edit1.clear()
                    self.line_edit2.clear()
                    self.table_widget.setItem(0, 1, QTableWidgetItem(''))
                    self.table_widget.setItem(0, 2, QTableWidgetItem(''))
                elif self.layer_name == 'Protected_Area':
                    self.line_edit3.clear()
                    self.line_edit4.clear()
                    self.table_widget.setItem(0, 1, QTableWidgetItem(''))
                    self.table_widget.setItem(0, 2, QTableWidgetItem(''))

            # --- Read Excel file (.xlsx or .xls) ---
            ext = os.path.splitext(path)[-1].lower()
            # Read data from Excel depending on file extension
            if ext == '.xlsx':
                book = openpyxl.load_workbook(path)
                sheet = book.active
                data = [
                    [sheet.cell(row=r + 1, column=c + 1).value for c in range(
                        sheet.max_column)]
                    for r in range(sheet.max_row)
                    if any(
                        sheet.cell(
                            row=r + 1, column=c + 1
                            ).value for c in range(sheet.max_column))
                ]
            elif ext == '.xls':
                book = xlrd.open_workbook(path)
                sheet = book.sheet_by_index(0)
                data = [
                    [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                    for r in range(sheet.nrows)
                    if any(sheet.cell_value(r, c) for c in range(sheet.ncols))
                ]
            else:
                return  # unsupported file

            # Check if there are at least 3 columns with information
            if any(len(row) < 3 or not all(row[i] for i in range(3)) for row in data):
                # Write a warning message and do not populate the table
                QMessageBox.information(
                    self, messages['warning_title'][lang],
                    messages['insufficient_data'][lang])
            else:
                # Check each cell in the first column to ensure it matches
                # the required format
                invalid_format = False
                for columnvalues in data:
                    value = columnvalues[0]  # Get value of column 1 (Excel)
                    if (not isinstance(value, str) or
                            not re.match(r'^(?:[NnSs]|[Հհ][սվ])\s\d{1,2}[A-Za-z]$',
                            value)):
                        invalid_format = True
                        break
                # If the format is not valid, write a warning message
                # and do not populate the table
                if invalid_format:
                    QMessageBox.information(
                        self, messages['warning_title'][lang],
                        messages['invalid_format'][lang])
                else:
                    max_row = len(data)
                    # Check if the number of rows is less than 3
                    if max_row < 3:
                        self.push_button_1.setDisabled(True)
                        self.push_button_2.setDisabled(True)
                    # Check if the number of rows is less than 4
                    if max_row < 4:
                        QMessageBox.information(
                            self, messages['warning_title'][lang],
                            messages['min_points_required'][lang])
                        # Clear all rows except the first one
                        self.clear_all_rows(
                            self.main_instance, self.table_widget,
                            self.line_edit1, self.line_edit2, self.line_edit3,
                            self.push_button_1, self.push_button_2,
                            self.push_button_3, self.push_button_4)
                    else:
                        # Populate the table with data starting from the second row
                        self.table_widget.setRowCount(max_row+1)
                        for row, columnvalues in enumerate(data, start=1):
                            for col, value in enumerate(columnvalues):
                                # Adjust column index to start from column 4
                                table_col = col + 3
                                item = QTableWidgetItem(str(value))
                                self.table_widget.setItem(row, table_col, item)
                        self.push_button_1.setEnabled(True)
                        self.push_button_2.setEnabled(True)
                        self.push_button_3.setEnabled(True)
                        self.push_button_4.setEnabled(True)

                        self.numbering(self.main_instance, self.table_widget)
                        self.distance_calc(
                            self.main_instance, self.table_widget)
                        if self.layer_name == 'Monument_Area':
                            self.area_calc(
                                self.main_instance, 'Monument_Area',
                                self.line_edit1, self.line_edit2,
                                self.line_edit3)
                            self.create_polygonMA(self.main_instance)
                        else:
                            self.area_calc(
                                self.main_instance, 'Protected_Area',
                                self.line_edit3, self.line_edit4,
                                self.line_edit1)
                            self.create_polygonPA(self.main_instance)
        else:
            pass
        self.close()
        self.main_instance.pantool()
